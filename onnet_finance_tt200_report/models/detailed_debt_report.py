# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
import logging

from odoo import models, _, fields, api
from odoo.tools import get_lang
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import io
import os
from datetime import datetime
from collections import defaultdict
from odoo.exceptions import UserError
_logger = logging.getLogger(__name__)


class DetailedDebtReportCustomHandler(models.AbstractModel):
    _name = 'detailed.debt.report.custom.handler'

    _inherit = 'account.general.ledger.report.handler'
    _description = 'Cash Book Report Custom Handler'

    def _get_row_json_data(self, data):

        result = {}
        row = data.get('columns')
        for cell in row:
            if cell.get('expression_label'):
                result[cell['expression_label']] = cell.get('no_format')

        if result.get('invoice_date') not in [None, '']:
            result['invoice_date'] = result.get('invoice_date').strftime("%d-%m-%Y")
            result['date'] = result['invoice_date']
            del result['invoice_date']
        if result.get('name') not in [None, '']:
            result['line_name'] = result['name']
        else:
            result['line_name'] = data['name']

        return result

    def _compact_and_expand_children_for_line(self, line, options):
        options['export_mode'] = 'excel'
        result = []
        data_input = [line]
        if 'account.account' in line.get('id') and 'expand_function' in line:
            data_input += self.env.ref('onnet_finance_tt200_report.detailed_debt_report').get_expanded_lines(
                options, line.get('id'), line.get('group_by'), line.get('expand_function'), line.get('progress'), line.get('offset', 0))
        for data in data_input:
            line_value = self._get_row_json_data(data)
            line_value['line_name'] = "Số dư đầu kỳ" if 'Initial Balance' in line_value['line_name'] else 'Tổng cộng' if 'Total' in line_value.get(
                'line_name') else line_value['line_name']
            line_value['parent_id'] = data.get('parent_id')
            line_value['id'] = data.get('id')
            line_value['level'] = len(line_value['line_name'].split()[0]) - \
                3 if line_value['line_name'].split()[0].isdigit() else -1
            if line_value.get('parent_id') is None:
                line_value['parent_id'] = line.get('id')
            if line_value.get('line_name') in ['Tổng cộng']:
                line_value['level'] = 1

            result.append(line_value)

        return result

    def _report_expand_unfoldable_line_general_ledger(self, line_dict_id, groupby, options, progress, offset, unfold_all_batch_data=None):
        def init_load_more_progress(line_dict):
            return {
                column['column_group_key']: line_col.get('no_format', 0)
                for column, line_col in zip(options['columns'], line_dict['columns'])
                if column['expression_label'] == 'balance'
            }

        report = self.env.ref('onnet_finance_tt200_report.detailed_debt_report')
        model, model_id = report._get_model_info_from_id(line_dict_id)

        if model != 'account.account':
            raise UserError(_("Wrong ID for general ledger line to expand: %s", line_dict_id))

        lines = []

        # Get initial balance
        if offset == 0:
            if unfold_all_batch_data:
                account, init_balance_by_col_group = unfold_all_batch_data['initial_balances'][model_id]
            else:
                account, init_balance_by_col_group = self._get_initial_balance_values(report, [model_id], options)[
                    model_id]

            initial_balance_line = report._get_partner_and_general_ledger_initial_balance_line(
                options, line_dict_id, init_balance_by_col_group, account.currency_id)

            if initial_balance_line:
                lines.append(initial_balance_line)

                # For the first expansion of the line, the initial balance line gives the progress
                progress = init_load_more_progress(initial_balance_line)

        # Get move lines
        if options['export_mode'] == 'excel':
            limit_to_load = None
        else:
            limit_to_load = report.load_more_limit + \
                1 if report.load_more_limit and options['export_mode'] != 'print' else None
        aml_results, has_more = self._get_aml_values(
            report, options, [model_id], offset=offset, limit=limit_to_load)
        aml_results = aml_results[model_id]

        next_progress = progress
        for aml_result in aml_results.values():
            new_line = self._get_aml_line(report, line_dict_id, options, aml_result, next_progress)
            lines.append(new_line)
            next_progress = init_load_more_progress(new_line)
        return {
            'lines': self._compute_line_number(lines),
            'offset_increment': report.load_more_limit,
            'has_more': has_more,
            'progress': next_progress,
        }

    def _compute_line_number(self, lines):
        index_code = 0
        index_number = 2
        map_number = {}
        result = []
        for line in lines:
            if line['columns'][index_code]:
                if line['columns'][index_code]['no_format'] not in map_number:
                    map_number[line['columns'][index_code]['no_format']] = len(map_number) + 1
                line['columns'][index_number]['name'] = map_number[line['columns'][index_code]['no_format']]
                line['columns'][index_number]['no_format'] = map_number[line['columns'][index_code]['no_format']]
            result.append(line)
        return result

    def export_to_xlsx(self, options, response=None):
        # super().ensure_one()
        report_id = self.env.ref('onnet_finance_tt200_report.detailed_debt_report')

        report_lines = report_id.get_report_information(options)['lines']
        data = {}
        data_list = []
        for line in report_lines:
            data_list += self._compact_and_expand_children_for_line(line, options)
        visited = []
        for line in data_list:
            if line in visited:
                continue
            visited.append(line)
            if line.get('parent_id'):
                id_split = line.get('id').split('|')
                parent_id_expected = '|'.join(id_split[:len(id_split) - 1])
                if data.get(parent_id_expected):
                    line['parent_id'] = parent_id_expected
            if line.get('line_name').split()[0].isdigit():
                del line['parent_id']
            if data.get(line.get('parent_id')):
                if not data.get(line['id'], {}) or line not in data.get(line['id'], {}).get('children'):
                    data[line['parent_id']]['children'].append(line)
            else:
                if line.get('children', [...]) == [...]:
                    line['children'] = []
                data[line['id']] = line
                data[line['id']]['children'].append(line)
        report_data = {}
        for d in data:
            row = data[d]
            report_data[row['line_name']] = row['children']

        return self.action_export_to_xlsx(report_data, options['date'])

    def _get_query_amls(self, report, options, expanded_account_ids, offset=0, limit=None):
        """ Construct a query retrieving the account.move.lines when expanding a report line with or without the load
        more.
        :param options:               The report options.
        :param expanded_account_ids:  The account.account ids corresponding to consider. If None, match every account.
        :param offset:                The offset of the query (used by the load more).
        :param limit:                 The limit of the query (used by the load more).
        :return:                      (query, params)
        """
        additional_domain = [('account_id', 'in', expanded_account_ids)] if expanded_account_ids is not None else None
        queries = []
        all_params = []
        lang = self.env.user.lang or get_lang(self.env).code
        journal_name = f"COALESCE(journal.name->>'{lang}', journal.name->>'en_US')" if \
            self.pool['account.journal'].name.translate else 'journal.name'
        account_name = f"COALESCE(account.name->>'{lang}', account.name->>'en_US')" if \
            self.pool['account.account'].name.translate else 'account.name'
        for column_group_key, group_options in report._split_options_per_column_group(options).items():
            # Get sums for the account move lines.
            # period: [('date' <= options['date_to']), ('date', '>=', options['date_from'])]
            tables, where_clause, where_params = report._query_get(
                group_options, domain=additional_domain, date_scope='strict_range')
            ct_query = report._get_query_currency_table(group_options)
            query = f'''
                (SELECT
                    account_move_line.id,
                    account_move_line.date,
                    account_move_line.date_maturity,
                    account_move_line.name,
                    account_move_line.ref,
                    account_move_line.company_id,
                    account_move_line.account_id,
                    account_move_line.payment_id,
                    account_move_line.partner_id,
                    account_move_line.currency_id,
                    account_move_line.amount_currency,
                    COALESCE(account_move_line.invoice_date, account_move_line.date)                 AS invoice_date,
                    ROUND(account_move_line.debit * currency_table.rate, currency_table.precision)   AS debit,
                    ROUND(account_move_line.credit * currency_table.rate, currency_table.precision)  AS credit,
                    ROUND(account_move_line.balance * currency_table.rate, currency_table.precision) AS balance,
                    move.name                               AS move_name,
                    company.currency_id                     AS company_currency_id,
                    partner.name                            AS partner_name,
                    move.move_type                          AS move_type,
                    account.code                            AS account_code,
                    offset_account                          AS offset_account,
                    {account_name}                          AS account_name,
                    journal.code                            AS journal_code,
                    {journal_name}                          AS journal_name,
                    full_rec.id                             AS full_rec_name,
                    %s                                      AS column_group_key
                FROM {tables}
                JOIN account_move move                      ON move.id = account_move_line.move_id
                LEFT JOIN {ct_query}                        ON currency_table.company_id = account_move_line.company_id
                LEFT JOIN res_company company               ON company.id = account_move_line.company_id
                LEFT JOIN res_partner partner               ON partner.id = account_move_line.partner_id
                LEFT JOIN account_account account           ON account.id = account_move_line.account_id
                LEFT JOIN account_journal journal           ON journal.id = account_move_line.journal_id
                LEFT JOIN account_full_reconcile full_rec   ON full_rec.id = account_move_line.full_reconcile_id
                WHERE {where_clause}
                ORDER BY account_move_line.date, account_move_line.id)
            '''

            queries.append(query)
            all_params.append(column_group_key)
            all_params += where_params
        full_query = " UNION ALL ".join(queries)

        if offset:
            full_query += ' OFFSET %s '
            all_params.append(offset)
        if limit:
            full_query += ' LIMIT %s '
            all_params.append(limit)

        return (full_query, all_params)

    def _dynamic_lines_generator(self, report, options, all_column_groups_expression_totals, warnings=None):
        lines = []
        date_from = fields.Date.from_string(options['date']['date_from'])
        company_currency = self.env.company.currency_id

        totals_by_column_group = defaultdict(lambda: {'debit': 0, 'credit': 0, 'balance': 0})
        for account, column_group_results in self._query_values(report, options):
            if not account.follow_by_partner:
                continue
            eval_dict = {}
            has_lines = False
            for column_group_key, results in column_group_results.items():
                account_sum = results.get('sum', {})
                account_un_earn = results.get('unaffected_earnings', {})

                account_debit = account_sum.get('debit', 0.0) + account_un_earn.get('debit', 0.0)
                account_credit = account_sum.get('credit', 0.0) + account_un_earn.get('credit', 0.0)
                account_balance = account_sum.get('balance', 0.0) + account_un_earn.get('balance', 0.0)

                eval_dict[column_group_key] = {
                    'amount_currency': account_sum.get('amount_currency', 0.0) + account_un_earn.get('amount_currency', 0.0),
                    'debit': account_debit,
                    'credit': account_credit,
                    'balance': account_balance,
                }

                max_date = account_sum.get('max_date')
                has_lines = has_lines or (max_date and max_date >= date_from)

                totals_by_column_group[column_group_key]['debit'] += account_debit
                totals_by_column_group[column_group_key]['credit'] += account_credit
                totals_by_column_group[column_group_key]['balance'] += account_balance

            lines.append(self._get_account_title_line(report, options, account, has_lines, eval_dict))
        # Report total line.
        for totals in totals_by_column_group.values():
            totals['balance'] = company_currency.round(totals['balance'])

        # Tax Declaration lines.
        journal_options = report._get_options_journals(options)

        if len(options['column_groups']) == 1 and len(journal_options) == 1 and journal_options[0]['type'] in ('sale', 'purchase'):
            lines += self._tax_declaration_lines(report, options, journal_options[0]['type'])

        # Total line
        lines.append(self._get_total_line(report, options, totals_by_column_group))

        return [(0, line) for line in lines]

    def action_export_to_xlsx(self, report_lines, date_range):
        for i in ['date_from', 'date_to']:
            date_range[i] = datetime.strptime(date_range.get(i, '1990-01-01'), "%Y-%m-%d")
        custom_path = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
        template_file_path = f"{custom_path}/static/report_templates/detailed_debt_templates.xlsx"

        wb = load_workbook(template_file_path)
        split_sheets = {}
        for key, value in report_lines.items():
            sheet_key = key[0:3]
            if not sheet_key.isdigit():
                continue
            if not sheet_key in split_sheets:
                split_sheets[sheet_key] = {}
            split_sheets[sheet_key][key] = value
        first_sheet_name = 'Sổ chi tiết công nợ'
        multi_sheet = {}

        for key, value in split_sheets.items():
            sheet_to_duplicate = wb[first_sheet_name]
            # Create a copy of the sheet
            new_sheet = wb.copy_worksheet(sheet_to_duplicate)

            # Rename the new sheet (optional)
            new_sheet.title = f"Tài khoản {key}"
            multi_sheet[key] = {
                'sheet': new_sheet,
                'data': value
            }
        wb.remove(wb[first_sheet_name])

        for key, sheet_data in multi_sheet.items():
            sheet = sheet_data['sheet']
            data = sheet_data['data']
            self.write_data_for_sheet(sheet, data, date_range)

        # save and return file
        filename_url = (f"detailed_debt_"
                        f"from_{date_range.get('date_from').strftime('%d%m%Y')}_to_{date_range.get('date_to').strftime('%d%m%Y')}.xlsx")
        fp = io.BytesIO()
        wb.save(fp)
        return {
            'file_name': filename_url,
            'file_content': fp.getvalue(),
            'file_type': 'xlsx',
        }

    def write_data_for_sheet(self, sheet, report_lines, date_range):
        company = self.env.user.company_id
        company_address = ', '.join(
            i for i in [company.street, company.street2, company.city, company.state_id.name, company.country_id.name]
            if i)
        exclude_table_values = {
            'company_name': {
                'col': 2,  # col B1
                'row': 1,
                'value': f"{company.display_name}"
            },
            'company_address': {
                'col': 2,  # col B2
                'row': 2,
                'value': f"{company_address}"
            },
            'name': {
                'col': 2,  # col B3
                'row': 3,
                'value': "SỔ CHI TIẾT CÔNG NỢ",
            },
            'date': {
                'col': 2,  # col B4
                'row': 4,
                'value': f"Từ ngày {date_range.get('date_from').strftime('%d-%m-%Y')} đến ngày {date_range.get('date_to').strftime('%d-%m-%Y')}",
            },
            'account': {
                'col': 2,  # col B5
                'row': 5,
                'value': f"Tài khoản: {', '.join(key for key in report_lines if report_lines[key][0].get('level',-1) == 0)}",
            },
        }

        for key, row in exclude_table_values.items():
            cell = sheet.cell(row=row['row'], column=row['col'])
            cell.value = row['value']
            merge_add = row.get('merge_add')
            if merge_add:
                sheet.merge_cells(start_row=row['row'],
                               start_column=row['col'], end_row=row['row'], end_column=row['col'] + merge_add)

        # # write move line row
        table_row = 9  # start from row 9
        row_lst = ['move_name', 'date', 'number', 'line_name', 'offset_account', 'debit', 'credit', 'balance']
        black_thin = Side(border_style="thin", color="000000")

        grouped_account_data = report_lines
        for account_display_name, move_lines_data in grouped_account_data.items():
            # region write header of account
            # first_move_line=move_lines_data.pop(0)
            level = move_lines_data[0].get('level', -1)
            is_root = level > -1 and level < 3

            # region add initial balance line
            index_trial_balance = 0 if len(move_lines_data) <= 1 else 1
            first_move_line = move_lines_data[index_trial_balance]

            init_debit = int(first_move_line.get('debit', 0))
            init_credit = int(first_move_line.get('credit', 0))
            move_lines_data = move_lines_data
            # endregion initial balance

            # region add ending and sum balance line
            sum_debit_lines = int(sum([row.get('debit', 0) for row in move_lines_data[index_trial_balance + 1:]]))
            sum_credit_lines = int(sum([row.get('credit', 0) for row in move_lines_data[index_trial_balance + 1:]]))
            ending_debit = int(sum_debit_lines + init_debit)
            ending_credit = int(sum_credit_lines + init_credit)

            ending_lines = [
                {
                    "line_name": "Cộng số phát sinh",
                    "debit": sum_debit_lines,
                    "credit": sum_credit_lines,
                    "balance": sum_debit_lines - sum_credit_lines
                },
                {
                    "line_name": "Số dư Cuối kỳ",
                    "debit": ending_debit,
                    "credit": ending_credit,
                    "balance": ending_debit - ending_credit
                }
            ]

            if not is_root:
                move_lines_data = move_lines_data + ending_lines
            # endregion ending lines
            # endregion header

            # write account lines
            for index, row in enumerate(move_lines_data):
                for col_index, col_name in enumerate(row_lst):
                    col_index += 1
                    if index == 0 and col_name == 'line_name':
                        col_index = 1
                        sheet.merge_cells(start_row=table_row, start_column=2, end_row=table_row, end_column=5)
                    cell = sheet.cell(row=table_row, column=col_index + 1)
                    cell_value = row.get(col_name, '')
                    if col_name in ['debit', 'credit', 'balance']:
                        cell.value = (int(row.get(col_name, 0)))
                        cell.number_format = '#,##0'
                    else:
                        cell.value = cell_value
                    self.css_cell_text(cell, bold=row.get(row_lst[1]) is None)
                    if is_root:
                        cell.fill = PatternFill(start_color='FDF5E6', fill_type="solid")

                    # css for cell
                    cell.border = Border(top=black_thin, left=black_thin, right=black_thin, bottom=black_thin)

                # increase num and index of row
                table_row += 1

        # region sign
        sign_row = table_row + 2
        sign_rows = [

            {
                'value': 'Kế toán ghi sổ',  # Ax-Bx
                'row': sign_row,
                'col': 2,
                'bold': True,
                'merge_add': 3
            },
            {
                'value': '(Ký, họ tên)',  # Ax+1-Bx+1
                'row': sign_row + 1,
                'col': 2,
                'merge_add': 3
            },
            {
                'value': 'Kế toán trưởng',  # Ex
                'row': sign_row,
                'col': 7,
                'bold': True,
                'merge_add': 2
            },
            {
                'value': '(Ký, họ tên)',  # Ex+1
                'row': sign_row + 1,
                'col': 7,
                'merge_add': 2
            },
            {
                'value': 'Ngày ..... tháng ..... năm ..... ',  # Gx-Hx-Ix
                'row': sign_row - 1,
                'col': 7,
                'italic': True,
                'merge_add': 2,
            },

        ]

        # cell Cxx
        for sign in sign_rows:
            cell = sheet.cell(row=sign['row'], column=sign['col'])
            cell.font = Font(size=14)
            cell.value = sign['value']
            if not sign.get('skip_text_center'):
                self.css_cell_text(cell, text_center=True)
            if sign.get('bold'):
                self.css_cell_text(cell, bold=True)
            if sign.get('italic'):
                self.css_cell_text(cell, italic=True)

            merge_add = sign.get('merge_add')
            if merge_add:
                sheet.merge_cells(start_row=sign['row'],
                               start_column=sign['col'], end_row=sign['row'], end_column=sign['col'] + merge_add)

    def css_cell_text(self, cell, text_center=False, bold=False, italic=False, font_name='Times New Roman', size=11):
        cell.font = Font(font_name, size=size)
        if text_center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if bold:
            cell.font = Font(font_name, bold=True, size=size)
        if italic:
            cell.font = Font(font_name, italic=True, size=size)
