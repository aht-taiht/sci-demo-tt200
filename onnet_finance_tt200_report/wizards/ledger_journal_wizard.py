# -*- coding: utf-8 -*-

from odoo import models, api, fields, _
from odoo.exceptions import ValidationError, UserError
from openpyxl import load_workbook
from odoo.tools import get_lang
from openpyxl.styles import Alignment, Border, Side, Font
from itertools import groupby
from operator import itemgetter
import os
import io
import base64
import time


ENTRIES_STATES = [
    ('all_posted_entries', 'All Posted Entries'),
    ('all_entries', 'All Entries'),
]


class LedgerJournalReportWizard(models.TransientModel):
    _name = 'ledger.journal.report.wizard'

    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date', default=fields.Date.today())
    entry_status = fields.Selection(ENTRIES_STATES, string='Status Entries', default='all_posted_entries',
                                    help="All Posted Entries: Get all account move line with state as posted\n"
                                         "All Entries: With state for both posted and draft")
    account_ids = fields.Many2many('account.account', string="Accounts")
    company_id = fields.Many2one('res.company', string="Company", default=lambda self: self.env.company)
    excel_file = fields.Binary('Report file ')
    file_name = fields.Char('Excel file', size=64)

    @api.constrains('date_from', 'date_to')
    def check_constrains_fields(self):
        for wizard in self:
            if wizard.date_from and wizard.date_from > (wizard.date_to or fields.Date.today()):
                raise UserError(_('Start Date must be not great than End Date!'))

    def action_export_to_xlsx(self):
        self.ensure_one()
        # load template file
        custom_path = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
        template_file_path = f"{custom_path}/static/report_templates/ledger_journal_template.xlsx"

        # region write data
        company = self.company_id
        company_address = ', '.join(
            i for i in [company.street, company.street2, company.city, company.state_id.name, company.country_id.name]
            if i)

        wb = load_workbook(template_file_path)
        ws = wb.active

        # write data not in table
        exclude_table_values = {
            'company_name': {
                'col': 1,  # col A1
                'row': 1,
                'value': f"Đơn vị: {company.display_name}"
            },
            'company_address': {
                'col': 1,  # col B1
                'row': 2,
                'value': f"Địa chỉ: {company_address}"
            },
            'date': {
                'col': 3,  # col C7
                'row': 7,
                'value': f"Từ ngày {self.date_from.strftime('%d/%m/%Y')} đến ngày {self.date_to.strftime('%d/%m/%Y')}",
                'merge_add': 2
            },
            'account': {
                'col': 3,  # col C8
                'row': 8,
                'value': f"Tài khoản: {','.join([account.code for account in self.account_ids])}",
                'merge_add': 2
            },
            'currency': {
                'col': 3,  # col C9
                'row': 9,
                'value': f"Tiền tệ: {company.currency_id.name}",
                'merge_add': 2
            },
        }
        for key, row in exclude_table_values.items():
            cell = ws.cell(row=row['row'], column=row['col'])
            cell.value = row['value']
            merge_add = row.get('merge_add')
            if merge_add:
                ws.merge_cells(start_row=row['row'],
                               start_column=row['col'], end_row=row['row'], end_column=row['col'] + merge_add)

        # write move line row
        table_row = 13  # start from row 13
        row_lst = ['date', 'move_name', 'date', 'line_name', 'offset_account', 'debit', 'credit']
        black_thin = Side(border_style="thin", color="000000")

        grouped_account_data = self.move_lines_data()
        for account_display_name, move_lines_data in grouped_account_data.items():
            # region write header of account
            cell = ws.cell(row=table_row, column=1)
            cell.value = account_display_name
            self.css_cell_text(cell, bold=True)
            ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=7)
            table_row += 1

            # region add initial balance line
            first_move_line = move_lines_data[0]
            init_debit = int(first_move_line.get('init_debit', 0))
            init_credit = int(first_move_line.get('init_credit', 0))
            initial_line = [
                {
                    "line_name": "Số dư Đầu kỳ",
                    "debit": init_debit,
                    "credit": init_credit,
                }
            ]
            move_lines_data = initial_line + move_lines_data
            # endregion initial balance

            # region add ending and sum balance line
            sum_debit_lines = int(sum([row['debit'] for row in move_lines_data[1:]]))
            sum_credit_lines = int(sum([row['credit'] for row in move_lines_data[1:]]))

            ending_debit = int(sum_debit_lines + init_debit)
            ending_credit = int(sum_credit_lines + init_credit)

            ending_lines = [
                {
                    "line_name": "Cộng số phát sinh",
                    "debit": sum_debit_lines,
                    "credit": sum_credit_lines,
                },
                {
                    "line_name": "Số dư Cuối kỳ",
                    "debit": ending_debit,
                    "credit": ending_credit
                }
            ]
            move_lines_data = move_lines_data + ending_lines
            # endregion ending lines
            # endregion header

            # write account lines
            for row in move_lines_data:
                for col_index, col_name in enumerate(row_lst):
                    cell = ws.cell(row=table_row, column=col_index + 1)
                    cell_value = row.get(col_name, '')
                    if col_name in ['debit', 'credit']:
                        cell.value = '{:,}'.format(int(row.get(col_name, 0)))
                    else:
                        cell.value = cell_value
                    # css for cell
                    if (col_index + 1) != 4 and (col_index + 1) not in [6, 7]:
                        self.css_cell_text(cell, text_center=True)
                    cell.border = Border(top=black_thin, left=black_thin, right=black_thin, bottom=black_thin)

                # increase num and index of row
                table_row += 1

        # region sign
        journal_row = table_row + 2
        sign_row = journal_row + 2
        sign_rows = [
            {
                'value': f"- Sổ này có ... trang, đánh số từ trang 01 đến trang ... ",  # Ax-Bx
                'row': journal_row - 1,
                'col': 1,
                'merge_add': 1,
                'skip_text_center': True,
            },
            {
                'value': f"- Ngày mở sổ {self.date_from.strftime('%d/%m/%Y')}",  # Ax-Bx
                'row': journal_row,
                'col': 1,
                'skip_text_center': True,
            },
            {
                'value': 'Người ghi sổ',  # Ax-Bx
                'row': sign_row,
                'col': 1,
                'bold': True,
                'merge_add': 1
            },
            {
                'value': '(Chữ ký, Họ và Tên)',  # Ax+1-Bx+1
                'row': sign_row + 1,
                'col': 1,
                'merge_add': 1
            },
            {
                'value': 'Kế toán trưởng',  # Ex
                'row': sign_row,
                'col': 3,
                'bold': True,
                'merge_add': 1
            },
            {
                'value': '(Chữ ký, Họ và Tên)',  # Ex+1
                'row': sign_row + 1,
                'col': 3,
                'merge_add': 1
            },
            {
                'value': '(Ngày...tháng...năm...)',  # Gx-Hx-Ix
                'row': sign_row - 1,
                'col': 5,
                'italic': True,
                'merge_add': 2,
            },
            {
                'value': 'Giám đốc',  # Gx-Hx-Ix
                'row': sign_row,
                'col': 5,
                'bold': True,
                'merge_add': 2,
            },
            {
                'value': '(Chữ ký, Họ và Tên, Dấu chức danh)',  # Gx+1-Hx+1-Ix+1
                'row': sign_row + 1,
                'col': 5,
                'merge_add': 2,
            },
        ]

        # cell Cxx
        for sign in sign_rows:
            cell = ws.cell(row=sign['row'], column=sign['col'])
            cell.font = Font(size=14)
            cell.value = sign['value']
            if not sign.get('skip_text_center'):
                self.css_cell_text(cell, text_center=True)
            if sign.get('bold'):
                self.css_cell_text(cell, bold=True)
            if sign.get('italic'):
                self.css_cell_text(cell, italic=True)

            merge_add = sign.get('merge_add')
            if merge_add:
                ws.merge_cells(start_row=sign['row'],
                               start_column=sign['col'], end_row=sign['row'], end_column=sign['col'] + merge_add)
        # endregion sign
        # endregion

        # save and return file
        current_ts = int(time.time())
        filename = f'ledger_journal_{current_ts}.xlsx'
        filename_url = (f"ledger_journal_"
                         f"from_{self.date_from.strftime('%d%m%Y')}_to_{self.date_to.strftime('%d%m%Y')}.xlsx")
        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.sudo().create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})
        return {
            'type': 'ir.actions.act_url',
            'name': filename_url,
            'url': '/web/content/ledger.journal.report.wizard/%s/excel_file/%s?download=true' % (
                export_id.id, filename_url),
            'target': 'new',
        }

    def css_cell_text(self, cell, text_center=False, bold=False, italic=False):
        if text_center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if bold:
            cell.font = Font(bold=True, size=14)
        if italic:
            cell.font = Font(italic=True, size=14)

    def get_where_clause(self):
        where_clause = f"aml.date >= '{self.date_from}' and aml.date <= '{self.date_to}'"
        if self.entry_status == 'all_posted_entries':
            where_clause += f" and am.state = 'posted' and aml.company_id = {self.company_id.id}"
        if self.account_ids:
            where_clause += f" and aa.id in {tuple(self.account_ids.ids)}".replace(",)", ")")

        return where_clause

    def move_lines_data(self):
        if self.pool['account.account'].name.translate:
            lang = self.env.user.lang or get_lang(self.env).code
            aa_name = f"COALESCE(aa.name->>'{lang}', aa.name->>'en_US')"
        else:
            aa_name = 'aa.name'
        aml_name = 'aml.name'

        where_clause = self.get_where_clause()
        query = f"""
            with account_init_balance as (
                select 
                    sum(aml.balance) as init_balance,
                    sum(aml.debit) as init_debit,
                    sum(aml.credit) as init_credit,
                    aa.id as account_id
                from account_move_line aml 
                join account_move am on am.id = aml.move_id
                join account_account aa on aa.id = aml.account_id
                where aml.date < '{self.date_from}' and aml.company_id = {self.company_id.id}
                    and aml.parent_state = 'posted'
                GROUP BY 4
            )
            select {aml_name} as line_name,
                   am.name as move_name,
                   am.id as move_id,
                   {aa_name} as account_name,
                   aa.id as account_id,
                   (aa.code || ' ' || {aa_name})  as account_display_name,
                   aib.init_debit as init_debit,
                   aib.init_credit as init_credit,
                   aml.debit AS debit,
                   aml.credit AS credit,
                   aml.balance AS balance,
                   to_char(aml.date, 'dd/mm/YYYY') as date,
                   offset_account as offset_account
            from account_move_line aml
            join account_move am on am.id = aml.move_id
            join account_account aa on aa.id = aml.account_id
            left join account_init_balance aib on aib.account_id = aa.id
            where {where_clause}
            order by account_id asc, date asc, line_name asc
        """

        self.env.cr.execute(query)
        move_lines_data = self.env.cr.dictfetchall()

        # group the data by the key
        grouped_account_data = {}
        for key, group in groupby(move_lines_data, key=itemgetter('account_display_name')):
            grouped_account_data[key] = list(group)
        return grouped_account_data
