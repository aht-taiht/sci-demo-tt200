# -*- coding: utf-8 -*-

from odoo import models, api, fields, _
from odoo.exceptions import ValidationError, UserError
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from odoo.tools import get_lang
from openpyxl.styles import Alignment, Border, Side, Font
import os
import io
import base64
import time
import datetime
import requests

ENTRIES_STATES = [
    ('all_posted_entries', 'All Posted Entries'),
    ('all_entries', 'All Entries'),
]


class TrialJournalReportWizard(models.TransientModel):
    _name = 'trial.balance.report.wizard'

    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date', default=fields.Date.today())
    entry_status = fields.Selection(ENTRIES_STATES, string='Status Entries', default='all_posted_entries',
                                    help="All Posted Entries: Get all account move line with state as posted\n"
                                         "All Entries: With state for both posted and draft")
    journal_ids = fields.Many2many('account.journal', string='Journals')
    excel_file = fields.Binary('Report file ')
    file_name = fields.Char('Excel file', size=64)

    @api.constrains('date_from', 'date_to')
    def check_constrains_fields(self):
        for wizard in self:
            if wizard.date_from and wizard.date_from > (wizard.date_to or fields.Date.today()):
                raise UserError(_('Start Date must be not great than End Date!'))

    def action_export_to_xlsx(self):
        self.ensure_one()
        # load template file
        custom_path = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
        template_file_path = f"{custom_path}/static/report_templates/trial_balance_template.xlsx"

        # region write data
        company = self.env.company
        company_address = ', '.join(
            i for i in [company.street, company.street2, company.city, company.state_id.name, company.country_id.name]
            if i)

        wb = load_workbook(template_file_path)
        ws = wb.active

        # write data not in table
        exclude_table_values = {
            'company_name': {
                'col': 1,  # col A1
                'row': 1,
                'value': f"Công ty: {company.display_name}",
            },
            'company_address': {
                'col': 1,  # col A2
                'row': 2,
                'value': f"Địa chỉ: {company_address}"
            },
            'form_no': {
                'col': 6,  # col F1
                'row': 1,
                'value': 'Mẫu số: S06 - DN',
                'is_center': True
            },
            'info_part_1': {
                'col': 6,  # col F2
                'row': 2,
                'value': '(Ban hành theo Thông tư số 200/2014/TT-BTC',
                'is_center': True
            },
            'info_part_2': {
                'col': 6,  # col F3
                'row': 3,
                'value': 'Ngày 22/12/2014 của Bộ Tài chính)',
                'is_center': True
            },
            'title': {
                'col': 1,  # col A5
                'row': 5,
                'value': f"BẢNG CÂN ĐỐI SỐ PHÁT SINH",
                'is_center': True
            },
            'date': {
                'col': 1,  # col A6
                'row': 6,
                'value': f"Từ ngày {self.date_from.strftime('%d-%m-%Y')} đến ngày {self.date_to.strftime('%d-%m-%Y')}",
                'is_center': True
            },

        }

        for key, row in exclude_table_values.items():
            cell = ws.cell(row=row['row'], column=row['col'])
            cell.value = row['value']
            self.css_cell_text(cell, text_center=row.get('text_center', False))
        # write move line row
        table_row = 11  # start from row 11
        index_row = 1

        move_lines_data = self.get_account_report_lines()  # self.move_lines_data()
        row_lst = [
            'account_code', 'line_name', 'opening_debit', 'opening_credit', 'arising_debit',
            'arising_credit', 'closing_debit', 'closing_credit']
        # print(move_lines_data)
        grey_medium = Side(border_style="medium", color="A6A6A6")
        black_medium = Side(border_style="medium", color="000000")
        for row in move_lines_data:
            for col_index, col_name in enumerate(row_lst):
                cell = ws.cell(row=table_row, column=col_index + 1)
                cell.value = row.get(col_name, '')
                # css for cell
                if col_name in ['opening_credit', 'opening_debit',
                                'arising_credit', 'arising_debit', 'closing_credit', 'closing_debit']:
                    cell.number_format = '#,##0'
                if col_index < len(row_lst) - 1:
                    cell.border = Border(top=grey_medium, left=grey_medium, right=grey_medium, bottom=grey_medium)
                else:
                    cell.border = Border(top=grey_medium, left=grey_medium, right=black_medium, bottom=grey_medium)
                cell.font = Font(name='Times New Roman')
                if row.get('is_root', False):
                    self.css_cell_text(cell, bold=True, background_color='FDF5E6')
                else:
                    self.css_cell_text(cell, background_color='FFFFFF')
                if row.get('line_name', '') == 'Tổng cộng':
                    self.css_cell_text(cell, bold=True, background_color='FAFAD2')

            # increase num and index of row
            index_row += 1
            table_row += 1

        # journal date
        journal_row = table_row + 1
        open_journal_date = "Ngày: .............................."
        ws.merge_cells(start_row=journal_row, start_column=6, end_row=journal_row, end_column=8)
        cell = ws.cell(row=journal_row, column=6)
        cell.value = open_journal_date
        cell.font = Font(name='Times New Roman')

        # sign region
        sign_row = journal_row + 2
        sign_rows = [
            {
                'value': 'Người lập biểu',  # A33-B33
                'row': sign_row,
                'col': 1,
                'bold': True,
                'merge_add': 1
            },
            {
                'value': '(Ký, họ tên)',  # A34-B44
                'row': sign_row + 1,
                'col': 1,
                'merge_add': 1
            },
            {
                'value': 'Kế toán trưởng',  # C33
                'row': sign_row,
                'col': 3,
                'bold': True,
                'merge_add': 2
            },
            {
                'value': '(Ký, họ tên)',  # C34
                'row': sign_row + 1,
                'col': 3,
                'merge_add': 2
            },
            {
                'value': 'Người đại diện theo pháp luật',  # F33
                'row': sign_row,
                'col': 6,
                'bold': True,
                'merge_add': 2,
            },
            {
                'value': '(Ký, họ tên)',  # F34
                'row': sign_row + 1,
                'col': 6,
                'merge_add': 2,
            },
        ]

        for sign in sign_rows:
            cell = ws.cell(row=sign['row'], column=sign['col'])

            cell.font = Font(name='Times New Roman')
            self.css_cell_text(cell, text_center=True)
            if sign.get('bold'):
                self.css_cell_text(cell, bold=True)
            if sign.get('italic'):
                self.css_cell_text(cell, italic=True)

            merge_add = sign.get('merge_add')
            if merge_add:
                ws.merge_cells(start_row=sign['row'],
                               start_column=sign['col'], end_row=sign['row'], end_column=sign['col'] + merge_add)
            try:
                cell.value = sign['value']
            except:
                pass
        # # endregion

        # # save and return file
        current_ts = int(time.time())
        filename = f'trial_balance_{current_ts}.xlsx'
        filename_url = (f"trial_balance_"
                        f"from_{self.date_from.strftime('%d%m%Y')}_to_{self.date_to.strftime('%d%m%Y')}.xlsx")
        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.sudo().create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})

        return {
            'type': 'ir.actions.act_url',
            'name': filename_url,
            'url': '/web/content/trial.balance.report.wizard/%s/excel_file/%s?download=true' % (
                export_id.id, filename_url),
            'target': 'new',
        }

    def get_account_report_lines(self):
        # Get trial_balance_report
        account_report_id = self.env.ref('account_reports.trial_balance_report')
        # get  report_line of trial_balance_report
        report_lines = account_report_id.get_report_information(
            self.get_option_for_account_report(account_report_id))['lines']

        # Get some fields needed for the report
        col_amount_title = ['opening_debit','opening_credit',
                'arising_debit', 'arising_credit' , 'closing_debit','closing_credit']
        report_line_jsons = {}
        for line in report_lines:
            temp = {}
            columns = line.get('columns', [])
            for index, value in enumerate(columns):
                temp[col_amount_title[index]] = value['no_format']
            temp['account_code'], temp['line_name'] = [line['name'].split()[0]] + [' '.join(line['name'].split()[1:])]
            report_line_jsons[temp['account_code']] = temp

        account_dict = self.get_json_all_account_account()
        return self._aggregate_report_line_with_account_account(report_line_jsons, account_dict, col_amount_title)

    def css_cell_text(self, cell, text_center=False, bold=False, italic=False, background_color=False):
        if text_center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if bold:
            cell.font = Font(name='Times New Roman', bold=True)
        if italic:
            cell.font = Font(name='Times New Roman', italic=True)
        if background_color:
            cell.fill = PatternFill(start_color=background_color, fill_type="solid")

    def get_where_clause(self, date_from=None, date_to=None):
        if date_from:
            where_clause = f"aml.date >= '{date_from}' and aml.date <= '{date_to}'"
        else:
            where_clause = f"aml.date <= '{date_to}'"
        if self.entry_status == 'all_posted_entries':
            where_clause += f" and am.state = 'posted' and aml.company_id = {self.env.company.id}"

        where_clause += f"and aml.parent_state != 'cancel'"
        return where_clause

    def get_option_for_account_report(self, instance):
        '''
            Set up condition for query line of report
            instance: report [object]
        '''
        options = instance.get_options()
        options['date']['date_from'] = self.date_from
        options['date']['date_to'] = self.date_to
        if self.entry_status != 'all_posted_entries':
            options['all_entries'] = True
        return options

    def get_json_all_account_account(self):
        '''
            get all account_account and convert list json to 1 json
            example: convert [{'account_code':'111','line_name':'my_name'}] ->['111':{'account_code':'111','line_name':'my_name}]
        '''
        lang = self.env.user.lang or get_lang(self.env).code
        query = f"""
            select COALESCE(aa.name->>'{lang}', aa.name->>'en_US') as line_name,
            aa.code as account_code
            from account_account aa
            order by code
        """
        self.env.cr.execute(query)
        accounts = self.env.cr.dictfetchall()
        return {a['account_code']: a for a in accounts}

    def _aggregate_report_line_with_account_account(self, report_line, account_account, col_amount_title):
        '''
            Aggregate like:
            331 : value
            3311 : value
            33111 : value
            33112 : value
            3312 : value
            
            Input:
                - report_line:  line from Trial Balance 
                - account_account: record of account_account
                - col_amoun_title: title of column report: opening_credit, opening_debit,...
            
        '''
        result = []
        added = []
        for account_code in account_account.keys():
            account_code_clear_0 = self._clear_0(account_code)
            number_child = 0
            for key in account_account.keys():
                if key.startswith(account_code_clear_0):
                    number_child += 1
                    detail = report_line.get(key, None)
                    if detail is None:
                        continue
                    for a in col_amount_title:
                        if account_account[account_code].get(a, None) is None:
                            account_account[account_code][a] = 0
                        account_account[account_code][a] += detail[a]
            account_account[account_code]['account_code'] = account_code_clear_0
            account_account[account_code]['is_root'] = len(
                account_code_clear_0) == 3 or number_child > 1 or account_code != account_code_clear_0
            if self._is_not_non_all_amount(account_account[account_code], col_amount_title) and account_code_clear_0 not in added:
                result.append(account_account[account_code])
                added.append(account_code_clear_0)
        if 'Total' in report_line:
            report_line['Total']['account_code'] = ''
            report_line['Total']['is_root'] = True
            report_line['Total']['line_name'] = 'Tổng cộng'
            result.append(report_line['Total'])
        return result

    def _clear_0(self, value, distance=100):
        if not value.isdigit():
            return value
        return value.rstrip('0') if int(value) // int(value.rstrip('0')) > distance else value

    def _is_not_non_all_amount(self, instance, amount_title):
        for a in amount_title:
            if a not in instance or instance.get(a, 'None') is None:
                return False
        return True
