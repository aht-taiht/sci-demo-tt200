# -*- coding: utf-8 -*-

from odoo import models, api, fields, _
from odoo.exceptions import ValidationError, UserError
from openpyxl import load_workbook
from odoo.tools import get_lang
from openpyxl.styles import Alignment, Border, Side, Font
import os
import io
import base64
import time


ENTRIES_STATES = [
    ('all_posted_entries', 'All Posted Entries'),
    ('all_entries', 'All Entries'),
]


class GeneralJournalReportWizard(models.TransientModel):
    _name = 'general.journal.report.wizard'

    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date', default=fields.Date.today())
    entry_status = fields.Selection(ENTRIES_STATES, string='Status Entries', default='all_posted_entries',
                                    help="All Posted Entries: Get all account move line with state as posted\n"
                                         "All Entries: With state for both posted and draft")
    journal_ids = fields.Many2many('account.journal', string='Journals')
    excel_file = fields.Binary('Report file ')
    file_name = fields.Char('Excel file', size=64)

    @api.constrains('date_from', 'date_to')
    def check_constrains_fields(self):
        for wizard in self:
            if wizard.date_from and wizard.date_from > (wizard.date_to or fields.Date.today()):
                raise UserError(_('Start Date must be not great than End Date!'))

    def action_export_to_xlsx(self):
        self.ensure_one()
        # load template file
        custom_path = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
        template_file_path = f"{custom_path}/static/report_templates/general_journal_template.xlsx"

        # region write data
        company = self.env.company
        company_address = ', '.join(
            i for i in [company.street, company.street2, company.city, company.state_id.name, company.country_id.name]
            if i)

        wb = load_workbook(template_file_path)
        ws = wb.active

        # write data not in table
        exclude_table_values = {
            'company_name': {
                'col': 1,  # col A1
                'row': 1,
                'value': f"Đơn vị: {company.display_name}"
            },
            'company_address': {
                'col': 1,  # col B1
                'row': 2,
                'value': f"Địa chỉ: {company_address}"
            },
            'date': {
                'col': 5,  # col E5
                'row': 5,
                'value': f"Từ ngày {self.date_from.strftime('%d/%m/%Y')} đến ngày {self.date_to.strftime('%d/%m/%Y')}"
            },
        }
        for key, row in exclude_table_values.items():
            cell = ws.cell(row=row['row'], column=row['col'])
            cell.value = row['value']

        # write move line row
        table_row = 13  # start from row 13
        index_row = 1
        move_lines_data = self.move_lines_data()
        row_lst = [
            'row_index', 'date', 'move_name', 'date', 'line_name', 'account_code', 'offset_account', 'debit', 'credit']

        black_thin = Side(border_style="thin", color="000000")
        for row in move_lines_data:
            for col_index, col_name in enumerate(row_lst):
                cell = ws.cell(row=table_row, column=col_index + 1)
                if col_name == 'row_index':
                    cell.value = index_row
                else:
                    cell.value = row.get(col_name, '')

                # css for cell
                if (col_index + 1) != 5:
                    self.css_cell_text(cell, text_center=True)
                if col_name in ['debit', 'credit']:
                    cell.number_format = '#,##0'
                cell.border = Border(top=black_thin, left=black_thin, right=black_thin, bottom=black_thin)
                cell.font = Font(name='Times New Roman')

            # increase num and index of row
            index_row += 1
            table_row += 1

        # journal date
        journal_row = table_row + 1
        open_journal_date = f"- Ngày mở sổ {self.date_from.strftime('%d/%m/%Y')}"
        ws.merge_cells(start_row=journal_row, start_column=1, end_row=journal_row, end_column=2)
        cell = ws.cell(row=journal_row, column=1)
        cell.value = open_journal_date

        # sign region
        sign_row = journal_row + 2
        sign_rows = [
            {
                'value': 'Người lập biểu',  # C22-D22
                'row': sign_row,
                'col': 3,
                'bold': True,
                'merge_add': 1
            },
            {
                'value': '(Ký, họ tên)',  # C23-D23
                'row': sign_row + 1,
                'col': 3,
                'merge_add': 1
            },
            {
                'value': 'Kế toán trưởng',  # E22
                'row': sign_row,
                'col': 5,
                'bold': True
            },
            {
                'value': '(Ký, họ tên)',  # E23
                'row': sign_row + 1,
                'col': 5
            },
            {
                'value': '(Ngày...tháng...năm...)',  # G20-H20-I20
                'row': sign_row - 1,
                'col': 7,
                'italic': True,
                'merge_add': 2,
            },
            {
                'value': 'Người đại diện theo pháp luật',  # G22-H22-I22
                'row': sign_row,
                'col': 7,
                'bold': True,
                'merge_add': 2,
            },
            {
                'value': '(Ký, họ tên)',  # G23-H23-I23
                'row': sign_row + 1,
                'col': 7,
                'merge_add': 2,
            },
        ]

        # cell C22
        for sign in sign_rows:
            cell = ws.cell(row=sign['row'], column=sign['col'])
            cell.value = sign['value']

            cell.font = Font(name='Times New Roman')
            self.css_cell_text(cell, text_center=True)
            if sign.get('bold'):
                self.css_cell_text(cell, bold=True)
            if sign.get('italic'):
                self.css_cell_text(cell, italic=True)

            merge_add = sign.get('merge_add')
            if merge_add:
                ws.merge_cells(start_row=sign['row'],
                               start_column=sign['col'], end_row=sign['row'], end_column=sign['col'] + merge_add)
        # endregion

        # save and return file
        current_ts = int(time.time())
        filename = f'general_journal_{current_ts}.xlsx'
        filename_url = (f"general_journal_"
                         f"from_{self.date_from.strftime('%d%m%Y')}_to_{self.date_to.strftime('%d%m%Y')}.xlsx")
        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.sudo().create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})
        return {
            'type': 'ir.actions.act_url',
            'name': filename_url,
            'url': '/web/content/general.journal.report.wizard/%s/excel_file/%s?download=true' % (
                export_id.id, filename_url),
            'target': 'new',
        }

    def css_cell_text(self, cell, text_center=False, bold=False, italic=False):
        if text_center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if bold:
            cell.font = Font(name='Times New Roman', bold=True)
        if italic:
            cell.font = Font(name='Times New Roman', italic=True)

    def get_where_clause(self):
        where_clause = f"aml.date >= '{self.date_from}' and aml.date <= '{self.date_to}'"
        if self.entry_status == 'all_posted_entries':
            where_clause += f" and am.state = 'posted' and aml.company_id = {self.env.company.id}"
        if self.journal_ids:
            where_clause += f" and aml.journal_id in {tuple(self.journal_ids.ids)}".replace(",)", ")")

        return where_clause

    def move_lines_data(self):
        if self.pool['account.account'].name.translate:
            lang = self.env.user.lang or get_lang(self.env).code
            aa_name = f"COALESCE(aa.name->>'{lang}', aa.name->>'en_US')"
        else:
            aa_name = 'aa.name'
        aml_name = 'aml.name'

        where_clause = self.get_where_clause()
        query = f"""
            select {aml_name} as line_name,
                   am.name as move_name,
                   am.id as move_id,
                   {aa_name} as account_name,
                   aa.code as account_code,
                   aml.debit AS debit,
                   aml.credit AS credit,
                   aml.balance AS balance,
                   to_char(aml.date, 'dd/mm/YYYY') as date,
                   offset_account as offset_account
            from account_move_line aml
            join account_move am on am.id = aml.move_id
            join account_account aa on aa.id = aml.account_id
            where {where_clause}
            order by date asc
        """

        self.env.cr.execute(query)
        return self.env.cr.dictfetchall()
