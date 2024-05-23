# -*- coding: utf-8 -*-

import os
import io
import base64
import time

from odoo.tools import get_lang
from odoo import models, api, fields, _
from odoo.exceptions import ValidationError, UserError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font

import logging
_logger = logging.getLogger(__name__)


class InventoryActivitiesReportWizard(models.TransientModel):
    _name = 'inventory.activities.report.wizard'
    _description = 'Inventory Activities Report Wizard'

    date_from = fields.Date(string='Start Date')
    date_to = fields.Date(string='End Date')

    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    warehouse_ids = fields.Many2many('stock.warehouse', string="Warehouse")
    warehouse_location_ids = fields.Many2many('stock.location', compute='_compute_warehouse_location_ids')
    location_ids = fields.Many2many('stock.location', string="Location")
    product_category_id = fields.Many2one('product.category', string="Product Category")
    product_id = fields.Many2one('product.product', string="Product")

    excel_file = fields.Binary('Report file ')
    file_name = fields.Char('Excel file', size=64)

    @api.constrains('date_from', 'date_to')
    def check_constrains_fields(self):
        for wizard in self:
            if wizard.date_from and wizard.date_from > (wizard.date_to or fields.Date.today()):
                raise UserError(_('Start Date must be not great than End Date!'))

    @api.depends('warehouse_ids')
    def _compute_warehouse_location_ids(self):
        for record in self:
            if record.warehouse_ids:
                view_location_ids = self.warehouse_ids.mapped('view_location_id').ids
                warehouse_location_ids = self.env['stock.location'].search(
                    [('id', 'child_of', view_location_ids), ('usage', '=', 'internal')])
            else:
                warehouse_location_ids = self.env['stock.location'].search([('usage', '=', 'internal')])

            record.warehouse_location_ids = warehouse_location_ids

    def action_export_to_xlsx(self):
        self.ensure_one()
        # load template file
        custom_path = os.path.realpath(os.path.join(os.path.dirname(__file__), '..'))
        template_file_path = f"{custom_path}/static/report_templates/inventory_report_activitied_template.xlsx"

        company_name = self.company_id.name
        if self.warehouse_ids:
            warehouse_name = ', '.join(warehouse.name for warehouse in self.warehouse_ids)
        else:
            warehouse_name = 'All'

        wb = load_workbook(template_file_path)
        ws = wb.active

        # write some data not in table
        exclude_table_values = {
            'company_name': {
                'col': 2,  # col B1
                'row': 1,
                'value': f"{company_name}",
                'merge_add': 3,
            },
            'warehouse_name': {
                'col': 2,  # col B2
                'row': 2,
                'value': f"{warehouse_name}",
                'merge_add': 3,
            },
            'date': {
                'col': 1,  # col E5
                'row': 5,
                'value': f"Từ ngày: {self.date_from.strftime('%d/%m/%Y')} đến ngày: {self.date_to.strftime('%d/%m/%Y')}",
                'bold': True,
                'merge_add': 11
            },
        }

        for key, row in exclude_table_values.items():
            cell = ws.cell(row=row['row'], column=row['col'])
            cell.value = row['value']
            if key == 'date':
                self.css_cell_text(cell, text_center=True)
            else:
                self.css_cell_text(cell, text_left=True)

            if row.get('bold'):
                self.css_cell_text(cell, bold=True)

            merge_add = row.get('merge_add')
            if merge_add:
                ws.merge_cells(start_row=row['row'],
                               start_column=row['col'], end_row=row['row'], end_column=row['col'] + merge_add)

        # write line row
        start_data_row = 10
        table_row = 10  # start from row 10
        index_row = 1
        lines_data = self.get_lines()
        row_lst = [
            'row_index',
            'product_code', 'product_name', 'product_uom',
            'init_quantity', 'init_value',
            'in_quantity', 'in_value', 'out_quantity', 'out_value',
            'end_quantity', 'end_value'
        ]

        black_thin = Side(border_style="thin", color="000000")
        for key, row in lines_data.items():
            for col_index, col_name in enumerate(row_lst):
                cell = ws.cell(row=table_row, column=col_index + 1)
                if col_name == 'row_index':
                    cell.value = index_row
                    self.css_cell_text(cell, text_center=True)
                elif col_name in ['product_name']:
                    cell.value = row.get(col_name, '')
                    self.css_cell_text(cell, text_left=True)
                else:
                    cell.value = row.get(col_name, '')
                    self.css_cell_text(cell, text_center=True)
                if col_index > 3:
                    cell.number_format = '#,##0'

                cell.border = Border(top=black_thin, left=black_thin, right=black_thin, bottom=black_thin)

            # increase num and index of row
            index_row += 1
            table_row += 1
        end_data_row = table_row - 1

        # add line total
        if end_data_row > start_data_row:
            for col_index, col_name in enumerate(row_lst):
                cell = ws.cell(row=table_row, column=col_index + 1)
                if col_name == 'row_index':
                    cell.value = 'Tổng'
                    self.css_cell_text(cell, text_left=True)
                if col_index > 3:
                    col_char = chr(97 + col_index).upper()
                    cell.value = '=SUM({col}{start}:{col}{end})'.format(
                        col=col_char, start=start_data_row, end=end_data_row)
                    cell.number_format = '#,##0'
                    self.css_cell_text(cell, text_center=True)
                cell.border = Border(top=black_thin, left=black_thin, right=black_thin, bottom=black_thin)
            table_row += 1

        # export date
        date_row = table_row + 1
        today = fields.Date.context_today(self)
        export_date = f"Ngày {today.day} tháng {today.month} năm {today.year}"
        cell = ws.cell(row=date_row, column=9)
        cell.value = export_date
        self.css_cell_text(cell, italic=True, text_center=True)

        # sign region
        sign_row = date_row + 2
        sign_rows = [
            {
                'value': 'Kế toán',
                'row': sign_row,
                'col': 3,
                'bold': True
            },
            {
                'value': '(Ký, họ tên)',
                'row': sign_row + 1,
                'col': 3,
            },
            {
                'value': 'Người lập',
                'row': sign_row,
                'col': 9,
                'bold': True
            },
            {
                'value': '(Ký, họ tên)',
                'row': sign_row + 1,
                'col': 9
            }
        ]

        # cell C22
        for sign in sign_rows:
            cell = ws.cell(row=sign['row'], column=sign['col'])
            cell.value = sign['value']
            self.css_cell_text(cell, text_center=True)
            if sign.get('bold'):
                self.css_cell_text(cell, bold=True)
            if sign.get('italic'):
                self.css_cell_text(cell, italic=True)

            merge_add = sign.get('merge_add')
            if merge_add:
                ws.merge_cells(start_row=sign['row'],
                               start_column=sign['col'], end_row=sign['row'], end_column=sign['col'] + merge_add)
        # endregion

        # save and return file
        current_ts = int(time.time())
        filename = f'inventory_report_{current_ts}.xlsx'
        filedname_url = (f"inventory_report_"
                         f"from_{self.date_from.strftime('%d%m%Y')}_to_{self.date_to.strftime('%d%m%Y')}.xlsx")
        fp = io.BytesIO()
        wb.save(fp)
        export_id = self.sudo().create(
            {'excel_file': base64.encodebytes(fp.getvalue()), 'file_name': filename})

        return {
            'type': 'ir.actions.act_url',
            'name': filedname_url,
            'url': '/web/content/inventory.activities.report.wizard/%s/excel_file/%s?download=true' % (
                export_id.id, filedname_url),
            'target': 'new',
        }

    def css_cell_text(self, cell, text_center=False, bold=False, italic=False, text_left=False):

        if text_center:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        if text_left:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if bold:
            cell.font = Font(bold=True, name='Times New Roman')
        if italic:
            cell.font = Font(italic=True, name='Times New Roman')

    def get_where_clause(self, type='init'):
        where_clause = f"sm.state = 'done'"
        # if type == 'init':
        #     where_clause += f" and sm.date < '{self.date_from}'"
        # elif type == 'period':
        #     where_clause += f" and sm.date >= '{self.date_from}' and sm.date <= '{self.date_to}'"
        # else:
        #     where_clause += f" and sm.date <= '{self.date_to}'"

        # if self.company_id:
        #     where_clause += f" and sm.company_id = {self.company_id.id}"

        if type == 'init':
            where_clause += f" and svl.create_date < '{self.date_from}'"
        elif type == 'period':
            where_clause += f" and svl.create_date >= '{self.date_from}' and svl.create_date <= '{self.date_to}'"
        else:
            where_clause += f" and svl.create_date <= '{self.date_to}'"

        if self.company_id:
            where_clause += f" and svl.company_id = {self.company_id.id}"

        if self.warehouse_ids:
            where_clause += f" and sm.warehouse_id in {tuple(self.warehouse_ids.ids)}".replace(",)", ")")

        if self.location_ids:
            where_clause += f" and (sm.location_id in {tuple(self.location_ids.ids)} or sm.location_dest_id in {tuple(self.location_ids.ids)})".replace(",)", ")")
        else:
            where_clause += f" and (source_location.usage = 'internal' or  dest_location.usage = 'internal')"

        return where_clause

    def query_datas(self, type='init'):
        lang = self.env.user.lang or get_lang(self.env).code

        pt_name = 'pt.name'
        uom_name = 'uom.name'

        if self.pool['product.template'].name.translate:
            pt_name = f"COALESCE(pt.name->>'{lang}', pt.name->>'en_US')"

        if self.pool['uom.uom'].name.translate:
            uom_name = f"COALESCE(uom.name->>'{lang}', uom.name->>'en_US')"

        where_clause = self.get_where_clause(type)

        # date_type = ('1-init', 'Đầu kì'), ('2-during', 'Trong kỳ'), ('3-end', 'Cuối kỳ')

        # query = f"""
        #     SELECT
        #         pp.default_code as product_code,
        #         {pt_name} as product_name,
        #         {uom_name} as product_uom,
        #         (
        #             SELECT
        #                 sum(svl.quantity)
        #             FROM stock_valuation_layer svl
        #             WHERE svl.stock_move_id = sm.id
        #         ) as quantity,
        #         (
        #             SELECT
        #                 sum(svl.value)
        #             FROM stock_valuation_layer svl
        #             WHERE svl.stock_move_id = sm.id
        #         ) as value,
        #         source_location.id AS from_location_id,
        #         dest_location.id AS to_location_id,
        #         source_location.usage AS source_location_usage,
        #         dest_location.usage AS dest_location_usage
        #     FROM stock_move sm
        #     JOIN stock_location source_location ON source_location.id = sm.location_id
        #     JOIN stock_location dest_location ON dest_location.id = sm.location_dest_id
        #     JOIN uom_uom uom ON uom.id = sm.product_uom
        #     LEFT JOIN product_product pp ON pp.id = sm.product_id
        #     LEFT JOIN product_template pt ON pt.id = pp.product_tmpl_id
        #     WHERE {where_clause}
        #     ORDER BY date asc
        # """

        query = f"""
            SELECT
                pp.id as product_id,
                pp.default_code as product_code,
                {pt_name} as product_name,
                {uom_name} as product_uom,
                svl.quantity,
                svl.value,
                source_location.id AS from_location_id,
                dest_location.id AS to_location_id,
                source_location.usage AS source_location_usage,
                dest_location.usage AS dest_location_usage
            FROM stock_valuation_layer svl
            LEFT JOIN stock_move sm ON sm.id = svl.stock_move_id
            LEFT JOIN stock_location source_location ON source_location.id = sm.location_id
            LEFT JOIN stock_location dest_location ON dest_location.id = sm.location_dest_id
            LEFT JOIN product_product pp ON pp.id = sm.product_id
            LEFT JOIN product_template pt ON pt.id = pp.product_tmpl_id
            LEFT JOIN uom_uom uom ON uom.id = pt.uom_id

            WHERE {where_clause}
            ORDER BY svl.create_date asc
        """

        self.env.cr.execute(query)
        return self.env.cr.dictfetchall()

    def get_datas_init(self):
        return self.query_datas(type='init')

    def get_datas_period(self):
        return self.query_datas(type='period')

    def get_lines(self):
        total_lines = {}

        init_lines = self.get_datas_init()
        period_lines = self.get_datas_period()

        for line in init_lines:
            if str(line['product_id']) not in total_lines:
                total_lines.update({
                    str(line['product_id']): {
                        'product_code': line['product_code'],
                        'product_name': line['product_name'],
                        'product_uom': line['product_uom'],
                        'init_quantity': line['quantity'],
                        'init_value': line['value'],
                        'in_quantity': 0,
                        'in_value': 0,
                        'out_quantity': 0,
                        'out_value': 0,
                        'end_quantity': line['quantity'],
                        'end_value': line['value'],
                        'source_location_usage': line['source_location_usage'],
                        'dest_location_usage': line['dest_location_usage'],
                    }
                })
            else:
                total_lines[str(line['product_id'])].update({
                    'init_quantity': total_lines[str(line['product_id'])]['init_quantity'] + line['quantity'],
                    'init_value': total_lines[str(line['product_id'])]['init_value'] + line['value'],
                    'end_quantity': total_lines[str(line['product_id'])]['end_quantity'] + line['quantity'],
                    'end_value': total_lines[str(line['product_id'])]['end_value'] + line['value'],
                })

        for line in period_lines:
            if line['source_location_usage'] == 'internal' and line['dest_location_usage'] != 'internal':
                line_type = 'out'
            elif line['source_location_usage'] != 'internal' and line['dest_location_usage'] == 'internal':
                line_type = 'in'
            else:
                line_type = 'other'

            if line_type in ['in', 'out']:
                if str(line['product_id']) not in total_lines:
                    total_lines.update({
                        str(line['product_id']): {
                            'product_code': line['product_code'],
                            'product_name': line['product_name'],
                            'product_uom': line['product_uom'],
                            'init_quantity': 0,
                            'init_value': 0,
                            'in_quantity': line['quantity'] if line_type == 'in' else 0,
                            'in_value': line['value'] if line_type == 'in' else 0,
                            'out_quantity': line['quantity'] if line_type == 'out' else 0,
                            'out_value': line['value'] if line_type == 'out' else 0,
                            'end_quantity': line['quantity'] if line_type == 'in' else -line['quantity'],
                            'end_value': line['value'] if line_type == 'in' else -line['value'],
                            'source_location_usage': line['source_location_usage'],
                            'dest_location_usage': line['dest_location_usage'],
                        }
                    })
                else:
                    if line_type == 'in':
                        update_vals = {
                            'in_quantity': total_lines[str(line['product_id'])]['in_quantity'] + line['quantity'],
                            'in_value': total_lines[str(line['product_id'])]['in_value'] + line['value'],
                            'end_quantity': total_lines[str(line['product_id'])]['end_quantity'] + line['quantity'],
                            'end_value': total_lines[str(line['product_id'])]['end_value'] + line['value'],
                        }
                    else:
                        update_vals = {
                            'out_quantity': total_lines[str(line['product_id'])]['out_quantity'] + line['quantity'],
                            'out_value': total_lines[str(line['product_id'])]['out_value'] + line['value'],
                            'end_quantity': total_lines[str(line['product_id'])]['end_quantity'] - line['quantity'],
                            'end_value': total_lines[str(line['product_id'])]['end_value'] - line['value'],
                        }

                    total_lines[str(line['product_id'])].update(update_vals)

        return total_lines
